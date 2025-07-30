import os
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


def get_previous_month_sheet_name():
    """
    Retorna o nome da aba do mês anterior (ex: "FEV.25").
    """
    now = datetime.now()
    year, month = now.year, now.month
    if month == 1:
        prev_month, year = 12, year - 1
    else:
        prev_month = month - 1
    month_abbr = {1:"JAN",2:"FEV",3:"MAR",4:"ABR",5:"MAI",6:"JUN",
                  7:"JUL",8:"AGO",9:"SET",10:"OUT",11:"NOV",12:"DEZ"}
    return f"{month_abbr[prev_month]}.{str(year)[-2:]}"


def atualizar_planilha_original(caminho_planilha, caminho_csv):
    try:
        df_csv = pd.read_csv(caminho_csv)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar CSV:\n{e}")
        return

    sheet = get_previous_month_sheet_name()
    try:
        wb = load_workbook(caminho_planilha)
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir Excel:\n{e}")
        return

    if sheet not in wb.sheetnames:
        messagebox.showerror("Erro", f"Aba '{sheet}' não encontrada.")
        return
    ws = wb[sheet]

    # Mapeia cabeçalho
    header = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column+1)
              if ws.cell(row=1, column=i).value}
    for col in ["Nº NF","Chave de acesso","Autorização","Dt emissão"]:
        if col not in header:
            messagebox.showerror("Erro", f"Coluna '{col}' ausente em '{sheet}'")
            return

    for row in range(2, ws.max_row+1):
        nf = ws.cell(row=row, column=header["Nº NF"]).value
        if not nf:
            continue
        match = df_csv[df_csv["NF Completa"] == nf]
        if not match.empty:
            m = match.iloc[0]
            ws.cell(row=row, column=header["Chave de acesso"]).value = m["Chave"]
            ws.cell(row=row, column=header["Autorização"]).value = str(m["Protocolo"])
            ws.cell(row=row, column=header["Dt emissão"]).value = m["Data de Emissão"]
            ws.cell(row=row, column=header["Autorização"]).number_format = "@"

    try:
        wb.save(caminho_planilha)
        messagebox.showinfo("Sucesso", f"'{sheet}' atualizada com sucesso.")
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao salvar:\n{e}")


def selecionar_arquivos():
    caminho_xls = filedialog.askopenfilename(
        title="Selecione a planilha Excel",
        filetypes=[("Excel", "*.xlsx *.xls")]
    )
    if not caminho_xls:
        return

    pasta = os.path.dirname(caminho_xls)
    csv_padrao = os.path.join(pasta, "notas.csv")
    if not os.path.exists(csv_padrao):
        messagebox.showwarning(
            "Aviso",
            "notas.csv não encontrado. Selecione manualmente."
        )
        csv_padrao = filedialog.askopenfilename(
            title="Selecione notas.csv",
            filetypes=[("CSV", "*.csv")]
        )
        if not csv_padrao:
            return

    atualizar_planilha_original(caminho_xls, csv_padrao)


def run_planilhar_notas():
    root = tk.Tk()
    root.title("Atualizador de Planilhas")
    root.geometry("400x150")
    tk.Label(root, text="Selecione Excel e CSV").pack(pady=20)
    tk.Button(root, text="Abrir", command=selecionar_arquivos).pack()
    root.mainloop()